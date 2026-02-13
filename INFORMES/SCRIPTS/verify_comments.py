from openpyxl import load_workbook

# Cargar archivo generado
wb = load_workbook(r'C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\ENTREGABLES\12022026_RESUMENES\R_CAPTA_12022026_125151.xlsx')

print('=== HOJAS EN EL ARCHIVO ===')
print(wb.sheetnames)

print('\n=== COMENTARIOS EN RESUMEN_EJECUTIVO (Columna A - Indicadores) ===')
ws_exec = wb['Resumen_Ejecutivo']
for i in range(2, min(10, ws_exec.max_row + 1)):
    cell = ws_exec.cell(i, 1)
    indicator = cell.value
    comment = cell.comment.text if cell.comment else "SIN COMENTARIO"
    print(f"Fila {i}: {indicator} -> {comment}")

print('\n=== COMENTARIOS EN AGENTES (Fila 1 - Encabezados) ===')
ws_ag = wb['Agentes']
for i in range(1, min(8, ws_ag.max_column + 1)):
    cell = ws_ag.cell(1, i)
    header = cell.value
    comment = cell.comment.text if cell.comment else "SIN COMENTARIO"
    print(f"Col {i}: {header} -> {comment}")

print('\n=== COMENTARIOS EN RESUMEN_DIARIO (Fila 1 - Encabezados) ===')
ws_diario = wb['Resumen_Diario']
for i in range(1, min(8, ws_diario.max_column + 1)):
    cell = ws_diario.cell(1, i)
    header = cell.value
    comment = cell.comment.text if cell.comment else "SIN COMENTARIO"
    print(f"Col {i}: {header} -> {comment}")
