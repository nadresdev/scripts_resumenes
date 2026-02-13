import pandas as pd
import os
import glob
import re
from datetime import datetime

# Importar diccionario de comentarios
try:
    from col_comments import COL_COMMENTS
except ImportError:
    COL_COMMENTS = {}

def find_latest_file(directory, pattern='*.xlsx'):
    files = glob.glob(os.path.join(directory, pattern))
    # Filtrar temporales
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    
    if not files:
        return None
    latest_file = max(files, key=os.path.getmtime)
    return latest_file

def seconds_to_hms(seconds):
    try:
        if pd.isna(seconds): return "00:00:00"
        seconds = int(float(seconds))
        if seconds < 0: return "00:00:00"
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        return "{:02d}:{:02d}:{:02d}".format(h, m, s)
    except (ValueError, TypeError):
        return "00:00:00"

def format_percentage(val):
    try:
        if pd.isna(val): return "0.00 %"
        return "{:.2f} %".format(float(val))
    except:
        return str(val)

def format_float_2dec(val):
    try:
        if pd.isna(val): return 0.00
        return round(float(val), 2)
    except:
        return val

def generate_daily_summary():
    input_dir = r"C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\DETALLE_LEADS_UNICOS"
    output_dir = r"C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\RESUMEN_DIARIO"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"Buscando archivo DETALLE ms reciente en: {input_dir}")
    latest_file = find_latest_file(input_dir, pattern='*DETALLE_LEADS_UNICOS*.xlsx')
    
    if not latest_file:
        print("No se encontraron archivos DETALLE LEADS UNICOS vlidos.")
        return

    print(f"Procesando archivo: {latest_file}")
    
    try:
        original_sheets = pd.read_excel(latest_file, sheet_name=None)
        
        if 'Detalle_Leads_Unicos' in original_sheets:
            df = original_sheets['Detalle_Leads_Unicos']
        else:
             print("Error: No se encontr hoja Detalle_Leads_Unicos.")
             return
        
        # Filtrar por ao 2026 en adelante y procesar fechas
        if 'fxCreated' in df.columns:
            df['fxCreated'] = pd.to_datetime(df['fxCreated'], errors='coerce')
            df = df[df['fxCreated'].dt.year >= 2026].copy()
        else:
             print("Error: No se encontr columna fxCreated.")
             return

        if df.empty:
            print("No hay registros del 2026 en adelante.")
            return

        # === Clasificacin de Horario (Operativo, Extra, FDS) ===
        # Lunes=0, Domingo=6
        df['day_of_week'] = df['fxCreated'].dt.dayofweek
        df['hour'] = df['fxCreated'].dt.hour
        
        # Categorizar
        # FDS: Sabado (5) o Domingo (6)
        # Operativo: Lun-Vie (0-4) y hora >= 10 y < 18 (10:00 a 17:59)
        # Extra: Lun-Vie y fuera de rango operativo
        
        conditions = [
            (df['day_of_week'] >= 5), # FDS
            (df['hour'] >= 10) & (df['hour'] < 18) # Operativo (Lun-Vie implcito por el orden del primer if si FDS ya filtr)
             # Nota: El orden importa. Si es FDS toma primero. Si no es FDS, evalua hora.
        ]
        choices = ['FDS', 'OPERATIVO']
        # El resto ser EXTRA (Lun-Vie fuera de horario)
        df['time_category'] = np.select(conditions, choices, default='EXTRA')

        # --- Clculos Grupales ---
        df['fecha_group'] = df['fxCreated'].dt.date
        df['mes_sort'] = df['fxCreated'].dt.to_period('M')
        
        summary_daily = calculate_metrics(df.copy(), group_col='fecha_group')
        summary_daily.rename(columns={'fecha_group': 'fecha'}, inplace=True)
        summary_daily = summary_daily.sort_values('fecha', ascending=False)
        
        # Aadimos columna Mes para agrupar visualmente despus (o filtrar en loop)
        summary_daily['fecha'] = pd.to_datetime(summary_daily['fecha'])
        summary_daily['mes_sort'] = summary_daily['fecha'].dt.to_period('M')

        # --- Estructura Final Hbrida ---
        final_summary_list = []
        unique_months = summary_daily['mes_sort'].unique()
        unique_months = sorted(unique_months, reverse=True)
        
        for mes in unique_months:
            mes_label = mes.strftime('%B %Y').upper()
            separator_row = {col: '' for col in summary_daily.columns}
            separator_row['fecha'] = f"MES: {mes_label}"
            sep_df = pd.DataFrame([separator_row])
            
            df_month_source = df[df['fxCreated'].dt.to_period('M') == mes]
            df_month_source['dummy_group'] = 'Total Mes'
            month_stats = calculate_metrics(df_month_source.copy(), group_col='dummy_group')
            month_stats.rename(columns={'dummy_group': 'fecha'}, inplace=True)
            month_stats['fecha'] = "TOTAL MES"
            
            daily_data = summary_daily[summary_daily['mes_sort'] == mes].copy()
            daily_data['fecha'] = daily_data['fecha'].dt.strftime('%Y-%m-%d')
            
            final_summary_list.append(sep_df)
            final_summary_list.append(daily_data)
            final_summary_list.append(month_stats)
            
            empty_row = {col: '' for col in summary_daily.columns}
            final_summary_list.append(pd.DataFrame([empty_row]))
            
        final_df = pd.concat(final_summary_list, ignore_index=True)
        if 'mes_sort' in final_df.columns:
            final_df.drop(columns=['mes_sort'], inplace=True)

        # Actualizar hojas
        original_sheets['Resumen_Diario'] = final_df
        if 'Resumen_Mensual' in original_sheets:
            del original_sheets['Resumen_Mensual']
        
        # Obtener Provider del DF para mayor seguridad
        provider = "UNKNOWN"
        if 'provider' in df.columns:
             unique_provs = df['provider'].dropna().unique()
             if len(unique_provs) > 0:
                 raw_p = str(unique_provs[0]).upper().strip()
                 if 'CAPTA' in raw_p: provider = 'CAPTA'
                 elif 'PLAYFILM' in raw_p: provider = 'PLAYFILM'
                 elif 'STARTEND' in raw_p: provider = 'STARTEND'
                 else:
                     provider = "".join([c for c in raw_p if c.isalnum() or c in (' ', '_', '-')]).strip()
        
        # Fallback
        if provider == "UNKNOWN":
            base_name = os.path.basename(latest_file).upper()
            if 'CAPTA' in base_name: provider = 'CAPTA'
            elif 'PLAYFILM' in base_name: provider = 'PLAYFILM'
            elif 'STARTEND' in base_name: provider = 'STARTEND'
            else:
                parts = base_name.split('_')
                provider = parts[0] if len(parts) > 0 else "UNKNOWN"

        timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
        
        output_filename = f"{provider}_RESUMEN_DIARIO_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        print(f"Guardando Resumen Diario en: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df_sheet in original_sheets.items():
                if isinstance(df_sheet, pd.DataFrame):
                     df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
            
            wb = writer.book
            if 'Resumen_Diario' in wb.sheetnames:
                apply_comments(wb['Resumen_Diario'], final_df.columns, "Resumen_Diario")
            if 'Detalle_Leads_Unicos' in wb.sheetnames:
                 apply_comments(wb['Detalle_Leads_Unicos'], original_sheets['Detalle_Leads_Unicos'].columns, "Detalle_Leads_Unicos")

        print("Resumen generado exitosamente.")

    except Exception as e:
        print(f"Error generando Resumen Diario: {e}")
        import traceback
        traceback.print_exc()

import numpy as np

def calculate_metrics(df_grouper, group_col):
    if df_grouper.empty:
         # Retornar vaco con todas las cols
         cols = [group_col, 'leads_insertados', 'contactados', 'ventas', 'interacciones_total', 
                 'tiempo_total_llamadas_hms', 'contactabilidad_%', 'conversion_%',
                 'mediana_tmo_x_periodo_seg', 'mediana_tmo_x_periodo_hms', 
                 'mediana_tmo_venta_x_periodo_seg', 'mediana_tmo_venta_x_periodo_hms',
                 'mediana_sla_seg', 'mediana_sla_hms',
                 'sla_operativo_mediana_hms', 'sla_extra_mediana_hms', 'sla_fds_mediana_hms']
         return pd.DataFrame(columns=cols)

    grouped = df_grouper.groupby(group_col).size().reset_index(name='leads_insertados')
    contactados = df_grouper[df_grouper['contactado'] == True].groupby(group_col).size().reset_index(name='contactados')
    ventas = df_grouper[df_grouper['venta'] == True].groupby(group_col).size().reset_index(name='ventas')
    interacciones = df_grouper.groupby(group_col)['Interacciones_x_lead'].sum().reset_index(name='interacciones_total')
    
    time_cols = [f'timeCall{i}' for i in range(1, 11)]
    for col in time_cols:
         if col not in df_grouper.columns: df_grouper[col] = 0
         else: df_grouper[col] = pd.to_numeric(df_grouper[col], errors='coerce').fillna(0)
    
    df_grouper['total_seconds_call'] = df_grouper[time_cols].sum(axis=1)
    tiempo_total = df_grouper.groupby(group_col)['total_seconds_call'].sum().reset_index(name='tiempo_total_llamadas_seg')
    
    # --- Medianas Generales ---
    df_grouper['tmo_total_registro'] = pd.to_numeric(df_grouper['tmo_total_registro'], errors='coerce').fillna(0)
    tmo_mediana = df_grouper[df_grouper['tmo_total_registro'] > 0].groupby(group_col)['tmo_total_registro'].median().reset_index(name='mediana_tmo_x_periodo_seg')
    
    df_grouper['tmo_venta'] = pd.to_numeric(df_grouper['tmo_venta'], errors='coerce').fillna(0)
    tmo_venta_mediana = df_grouper[df_grouper['venta'] == True].groupby(group_col)['tmo_venta'].median().reset_index(name='mediana_tmo_venta_x_periodo_seg')
    
    df_grouper['sla_seg'] = pd.to_numeric(df_grouper['sla_seg'], errors='coerce')
    sla_mediana = df_grouper.groupby(group_col)['sla_seg'].median().reset_index(name='mediana_sla_seg')
    
    # --- Mediana ACW (timeAcw1...10 acumulado) ---
    acw_cols = [f'timeAcw{i}' for i in range(1, 11)]
    for col in acw_cols:
         if col not in df_grouper.columns: df_grouper[col] = 0
         else: df_grouper[col] = pd.to_numeric(df_grouper[col], errors='coerce').fillna(0)
    
    # Calcular ACW Total por Lead
    df_grouper['total_acw_lead'] = df_grouper[acw_cols].sum(axis=1)
    
    # Calcular Mediana Diaria de esos totales
    # (Opcional: Filtrar > 0 si solo interesa ACW efectivo, pero ACW puede ser 0 y vlido)
    # Asumimos que queremos la mediana de TODOS los leads gestionados (contactados?) o todos los insertados?
    # Usualmente se mide sobre los gestionados (contactados o con intento). 
    # Si un lead no se llam, su ACW es 0. Si se incluye en la mediana, baja mucho.
    # Filtraremos por contactado=True o tmo_total > 0 para ser consistente con TMO.
    acw_mediana = df_grouper[df_grouper['tmo_total_registro'] > 0].groupby(group_col)['total_acw_lead'].median().reset_index(name='timeAcw_mediana_dia_seg')

    
    # --- Medianas SLA por Franja (Operativo, Extra, FDS) ---
    # SLA Operativo
    df_op = df_grouper[df_grouper['time_category'] == 'OPERATIVO']
    if not df_op.empty:
        sla_op = df_op.groupby(group_col)['sla_seg'].median().reset_index(name='sla_operativo_mediana_seg')
    else:
        sla_op = pd.DataFrame(columns=[group_col, 'sla_operativo_mediana_seg'])

    # SLA Extra
    df_ex = df_grouper[df_grouper['time_category'] == 'EXTRA']
    if not df_ex.empty:
        sla_ex = df_ex.groupby(group_col)['sla_seg'].median().reset_index(name='sla_extra_mediana_seg')
    else:
        sla_ex = pd.DataFrame(columns=[group_col, 'sla_extra_mediana_seg'])
        
    # SLA FDS
    df_fds = df_grouper[df_grouper['time_category'] == 'FDS']
    if not df_fds.empty:
        sla_fds = df_fds.groupby(group_col)['sla_seg'].median().reset_index(name='sla_fds_mediana_seg')
    else:
        sla_fds = pd.DataFrame(columns=[group_col, 'sla_fds_mediana_seg'])

    # Merge Final
    summary = grouped
    summary = summary.merge(contactados, on=group_col, how='left').fillna({'contactados': 0})
    summary = summary.merge(ventas, on=group_col, how='left').fillna({'ventas': 0})
    summary = summary.merge(interacciones, on=group_col, how='left').fillna({'interacciones_total': 0})
    summary = summary.merge(tiempo_total, on=group_col, how='left').fillna({'tiempo_total_llamadas_seg': 0})
    summary = summary.merge(tmo_mediana, on=group_col, how='left')
    summary = summary.merge(tmo_venta_mediana, on=group_col, how='left')
    summary = summary.merge(sla_mediana, on=group_col, how='left')
    summary = summary.merge(acw_mediana, on=group_col, how='left')
    
    # Merge SLA Franjas
    summary = summary.merge(sla_op, on=group_col, how='left')
    summary = summary.merge(sla_ex, on=group_col, how='left')
    summary = summary.merge(sla_fds, on=group_col, how='left')
    
    # Calculos Finales
    summary['contactabilidad_%'] = (summary['contactados'] / summary['leads_insertados']) * 100
    summary['conversion_%'] = (summary['ventas'] / summary['leads_insertados']) * 100
    
    # Format %
    summary['contactabilidad_%'] = summary['contactabilidad_%'].apply(format_percentage)
    summary['conversion_%'] = summary['conversion_%'].apply(format_percentage)
    
    # Format Decimals
    float_cols = ['mediana_tmo_x_periodo_seg', 'mediana_tmo_venta_x_periodo_seg', 'mediana_sla_seg']
    for col in float_cols:
         if col in summary.columns: summary[col] = summary[col].apply(format_float_2dec)
         
    # Format HMS
    summary['tiempo_total_llamadas_hms'] = summary['tiempo_total_llamadas_seg'].apply(seconds_to_hms)
    summary['mediana_tmo_x_periodo_hms'] = summary['mediana_tmo_x_periodo_seg'].apply(seconds_to_hms)
    summary['mediana_tmo_venta_x_periodo_hms'] = summary['mediana_tmo_venta_x_periodo_seg'].apply(seconds_to_hms)
    summary['mediana_sla_hms'] = summary['mediana_sla_seg'].apply(seconds_to_hms)
    summary['timeAcw_mediana_dia_hms'] = summary['timeAcw_mediana_dia_seg'].apply(seconds_to_hms)
    
    # Format HMS Franjas
    summary['sla_operativo_mediana_hms'] = summary['sla_operativo_mediana_seg'].apply(seconds_to_hms)
    summary['sla_extra_mediana_hms'] = summary['sla_extra_mediana_seg'].apply(seconds_to_hms)
    summary['sla_fds_mediana_hms'] = summary['sla_fds_mediana_seg'].apply(seconds_to_hms)
    
    # Fill NA HMS
    hms_cols = ['tiempo_total_llamadas_hms', 'mediana_tmo_x_periodo_hms', 'mediana_tmo_venta_x_periodo_hms', 
                'mediana_sla_hms', 'sla_operativo_mediana_hms', 'sla_extra_mediana_hms', 'sla_fds_mediana_hms',
                'timeAcw_mediana_dia_hms']
    for col in hms_cols:
         summary[col] = summary[col].fillna("00:00:00")
    
    final_cols = [
        group_col, 'leads_insertados', 'contactados', 'ventas', 
        'interacciones_total', 'tiempo_total_llamadas_hms', 
        'contactabilidad_%', 'conversion_%',
        'mediana_tmo_x_periodo_seg', 'mediana_tmo_x_periodo_hms',
        'mediana_tmo_venta_x_periodo_seg', 'mediana_tmo_venta_x_periodo_hms',
        'mediana_sla_seg', 'mediana_sla_hms',
        'sla_operativo_mediana_hms', 'sla_extra_mediana_hms', 'sla_fds_mediana_hms',
        'timeAcw_mediana_dia_hms'
    ]
    return summary[final_cols]

def apply_comments(worksheet, columns, sheet_config_name):
    from openpyxl.comments import Comment
    if sheet_config_name not in COL_COMMENTS:
         found = False
         for k in COL_COMMENTS.keys():
             if k in sheet_config_name:
                 sheet_config_name = k
                 found = True
                 break
         if not found: return

    comments_dict = COL_COMMENTS[sheet_config_name]
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        comment_text = None
        if col_name in comments_dict:
            comment_text = comments_dict[col_name]
        else:
             clean_name = col_name.replace('periodo', 'dia').replace('total', 'dia')
             if clean_name in comments_dict:
                 comment_text = comments_dict[clean_name]

        if comment_text:
            cell.comment = Comment(comment_text, "System")

if __name__ == "__main__":
    generate_daily_summary()
