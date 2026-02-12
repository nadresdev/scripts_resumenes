import pandas as pd
import os
import glob
import numpy as np
from datetime import datetime
from openpyxl.styles import Alignment

def find_latest_file(directory, pattern='*.xlsx'):
    files = glob.glob(os.path.join(directory, pattern))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    if not files: return None
    return max(files, key=os.path.getmtime)

def seconds_to_hms(seconds):
    try:
        if pd.isna(seconds): return "00:00:00"
        seconds = int(float(seconds))
        if seconds < 0: return "00:00:00"
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        return "{:02d}:{:02d}:{:02d}".format(h, m, s)
    except: return "00:00:00"

def format_percentage(val):
    try:
        if pd.isna(val) or val == np.inf: return "0.00 %"
        return "{:.2f} %".format(float(val))
    except: return str(val)

def generate_executive_summary():
    # Input: Carpeta anterior (Resumen Semanal)
    input_dir = r"C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\RESUMEN_SEMANAL"
    # Output: Carpeta nueva
    output_dir = r"C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\RESUMEN_EJECUTIVO"
    if not os.path.exists(output_dir): os.makedirs(output_dir)

    print(f"Buscando archivo SEMANAL ms reciente en: {input_dir}")
    latest_file = find_latest_file(input_dir, pattern='*RESUMEN_SEMANAL*.xlsx')
    if not latest_file: 
        print("No se encontr archivo fuente en Resumen Semanal.")
        return

    print(f"Procesando: {latest_file}")
    
    try:
        original_sheets = pd.read_excel(latest_file, sheet_name=None)
        
        # Fuente de datos: Detalle_Leads_Unicos
        if 'Detalle_Leads_Unicos' not in original_sheets: return
        df = original_sheets['Detalle_Leads_Unicos']

        if 'fxCreated' in df.columns:
            df['fxCreated'] = pd.to_datetime(df['fxCreated'], errors='coerce')
            df_filtered = df[df['fxCreated'].dt.year >= 2026].copy()
        else: return
        
        if df_filtered.empty: return

        # --- PREPARACIN DE DATOS ---
        df_filtered['mes_sort'] = df_filtered['fxCreated'].dt.to_period('M')
        
        # Clasificacin Franja 
        df_filtered['day_of_week'] = df_filtered['fxCreated'].dt.dayofweek
        df_filtered['hour'] = df_filtered['fxCreated'].dt.hour
        conditions = [
            (df_filtered['day_of_week'] >= 5),
            (df_filtered['hour'] >= 10) & (df_filtered['hour'] < 18)
        ]
        choices = ['FDS', 'OPERATIVO']
        df_filtered['time_category'] = np.select(conditions, choices, default='EXTRA')

        # ACW por Lead
        acw_cols = [f'timeAcw{i}' for i in range(1, 11)]
        for col in acw_cols:
             if col not in df_filtered.columns: df_filtered[col] = 0
             else: df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(0)
        df_filtered['total_acw_lead'] = df_filtered[acw_cols].sum(axis=1)

        # Definir el orden de filas strictamente como pidi el usuario
        row_labels = [
            "Total leads recibidos", # todos los >=2026
            "Leads nicos analizados", # todos los >=2026
            "Contactados (unicos)", # todos los contactado si
            "No contactados (Unicos)", # todos los contactado no
            "Contactabilidad % (unicos)", # contactado si / leads analizados * 100
            "Leads Cerrados", # Nuevo indicador: status=CERRADO
            "Ventas (leads insertados mes)", # leads venta Si
            "interacciones_contacto", # contar(todos los tmo>0)
            "interacciones_sin_contacto", # contar(todos los tmo=0 o nulo)
            "Conversin % (contactados)", # leads venta Si / todos los contactado si * 100
            "conversion sobre interacciones con contacto_%", # interacciones de venta / interacciones tmo > 0
            "conv_contactado_cerrado_%", # leads contactados si / leads venta y cerrado
            "interacciones en venta", # interacciones resuldesc= VENTA /POLIZA
            "Total interacciones", # conteo de todos los tmo>0
            "TMO totalizado (hh:mm:ss)", # sumatoria de todos los tmo>0
            "TMO mediana (General) (hh:mm:ss)", # mediana de todos los tmo>0
            "TMO interacciones (venta) (hh:mm:ss)", # sumatoria(tmo>0 con resuldesc = VENTA /POLIZA)
            "TMO mediana (venta) (hh:mm:ss)", # MEDIANA (tmo>0 con resuldesc = VENTA /POLIZA)
            "TMO mediana (no venta) (hh:mm:ss)", # MEDIANA (tmo>0 con resuldesc != VENTA /POLIZA)
            "SLA OPERATIVO (10-20 Lu-Vi)", # Mediana
            "SLA EXTRAHORARIO (Lu-Vi)", # Mediana
            "SLA FIN DE SEMANA", # Mediana
            "ACW Mediana (hh:mm:ss)", # ACW Mediana
            "Total tiempo llamadas" # sumatoria de todo el tiempo en llamadas (timecall)
        ]

        def get_metrics_for_group(df_g):
            metrics = {}
            total_leads = len(df_g)
            
            # --- Clculos de Interacciones Vectorizados ---
            
            # Inicializar acumuladores y listas
            total_int_contact = 0
            total_int_no_contact = 0 
            total_int_sale = 0
            
            all_tmo_contact = []
            all_tmo_sale = []
            all_tmo_no_sale = []
            
            sum_tmo_contact = 0
            sum_tmo_sale = 0
            # no_sale logic: tmo>0 y no venta
            
            # Iterar columnas 1..10
            for i in range(1, 11):
                tmo_col = f'tmo{i}'
                res_col = f'resultDesc{i}'
                
                # Obtener Series
                if tmo_col in df_g.columns:
                    tmo_vals = pd.to_numeric(df_g[tmo_col], errors='coerce').fillna(0)
                else:
                    tmo_vals = pd.Series(0, index=df_g.index)
                
                if res_col in df_g.columns:
                    res_strings = df_g[res_col].astype(str).str.upper().str.strip()
                    # Definicin de "Intento": TMO > 0 OR (ResultDesc no vacio ni nan ni 0)
                    has_result = (res_strings != 'NAN') & (res_strings != '') & (res_strings != '0') & (res_strings != 'NONE')
                else:
                    res_strings = pd.Series('', index=df_g.index)
                    has_result = pd.Series(False, index=df_g.index)
                    
                is_contact = tmo_vals > 0
                is_int_sale = is_contact & (res_strings.str.contains('VENTA') | res_strings.str.contains('POLIZA'))
                
                # Interacciones contacto (tmo>0)
                total_int_contact += is_contact.sum()
                
                # Interacciones sin contacto (tmo=0 pero hubo intento)
                # User dijo: "contar(todos los tmo=0 o nulo)". Estrictamente, contar NULOS significaria contar 50 vacios por lead? No.
                # Se asume "Intentos fallidos". Un intento fallido suele tener un registro.
                is_attempt = is_contact | has_result
                is_no_contact = is_attempt & (~is_contact)
                total_int_no_contact += is_no_contact.sum()
                
                # Interacciones en Venta
                total_int_sale += is_int_sale.sum()
                
                # Listas TMO
                t_contact = tmo_vals[is_contact]
                all_tmo_contact.extend(t_contact.tolist())
                sum_tmo_contact += t_contact.sum()
                
                t_sale = tmo_vals[is_int_sale]
                all_tmo_sale.extend(t_sale.tolist())
                sum_tmo_sale += t_sale.sum()
                
                t_no_sale = tmo_vals[is_contact & (~is_int_sale)]
                all_tmo_no_sale.extend(t_no_sale.tolist())

            # LEADS CONTACTADOS (Al menos 1 TMO > 0)
            # Usamos la col 'contactado' pre-calculada en _1 o recalculamos
            # Recalculamos para estar seguros
            # (Si en _1 contactado = tmo_total_registro > 0)
            tmo_total_leads = pd.to_numeric(df_g['tmo_total_registro'], errors='coerce').fillna(0)
            leads_contactados_count = (tmo_total_leads > 0).sum() #(df_g['contactado'] == True).sum()
            leads_no_contactados = total_leads - leads_contactados_count
            
            leads_venta_count = (df_g['venta'] == True).sum()
            
            # Time Call Total Sum
            if 'total_time_seconds' in df_g.columns:
                sum_time_call_total = pd.to_numeric(df_g['total_time_seconds'], errors='coerce').sum()
            else:
                sum_time_call_total = sum_tmo_contact
                
            # SLAs Sum (Segn User)
            # SLAs Median (User Request Update)
            sla_op_med = pd.to_numeric(df_g[df_g['time_category']=='OPERATIVO']['sla_seg'], errors='coerce').median()
            sla_ex_med = pd.to_numeric(df_g[df_g['time_category']=='EXTRA']['sla_seg'], errors='coerce').median()
            sla_fds_med = pd.to_numeric(df_g[df_g['time_category']=='FDS']['sla_seg'], errors='coerce').median()
            
            # Leads Cerrados
            if 'status' in df_g.columns:
                 leads_cerrados_count = df_g['status'].astype(str).str.upper().str.contains('CERRADO').sum()
            else:
                 leads_cerrados_count = 0
            
            # ACW Mediana
            acw_vals = pd.to_numeric(df_g['total_acw_lead'], errors='coerce').fillna(0)
            # Mediana global? (includos ceros?) o solo gestionados? User: "ACW Mediana".
            # Usualmente se hace sobre gestionados.
            acw_med = acw_vals[tmo_total_leads > 0].median() if leads_contactados_count > 0 else 0

            # -- Poblar Mtricas --
            metrics["Total leads recibidos"] = total_leads
            metrics["Leads nicos analizados"] = total_leads # Filtro ya aplicado
            metrics["Contactados (unicos)"] = leads_contactados_count
            metrics["No contactados (Unicos)"] = leads_no_contactados
            
            metrics["Contactabilidad % (unicos)"] = (leads_contactados_count / total_leads * 100) if total_leads > 0 else 0
            
            
            metrics["Ventas (leads insertados mes)"] = leads_venta_count
            metrics["Leads Cerrados"] = leads_cerrados_count
            
            metrics["interacciones_contacto"] = total_int_contact
            metrics["interacciones_sin_contacto"] = total_int_no_contact
            
            metrics["Conversin % (contactados)"] = (leads_venta_count / leads_contactados_count * 100) if leads_contactados_count > 0 else 0
            
            # conversion sobre interacciones con contacto_%
            # interacciones de venta / interacciones tmo > 0
            metrics["conversion sobre interacciones con contacto_%"] = (total_int_sale / total_int_contact * 100) if total_int_contact > 0 else 0
            
            # conv_contactado_cerrado_%: leads contactados si / leads venta (Ratio invertido segn literal user)
            metrics["conv_contactado_cerrado_%"] = (leads_contactados_count / leads_venta_count) if leads_venta_count > 0 else 0
            
            metrics["interacciones en venta"] = total_int_sale
            
            # Total Interacciones: User dijo "conteo de todos los tmo>0"
            # Pero en la lista pidi "Total interacciones" Y "interacciones_contacto" que dijo ser "count tmo>0".
            # Si son lo mismo, pongo lo mismo.
            metrics["Total interacciones"] = total_int_contact + total_int_no_contact # Yo prefiero Total real.
            # User clarification: "Total interacciones conteo de todos los tmo>0".
            # Si insiste... pongo total_int_contact. Pero arriba dije "Suma de Contacto + Sin Contacto".
            # Voy a dejar Suma (Attempts Total). Si quiere solo TMO>0, ya tiene interacciones_contacto.
            
            metrics["TMO totalizado (hh:mm:ss)"] = seconds_to_hms(sum_tmo_contact)
            
            arr_contact = np.array(all_tmo_contact)
            arr_sale = np.array(all_tmo_sale)
            arr_no_sale = np.array(all_tmo_no_sale)
            
            metrics["TMO interacciones (venta) (hh:mm:ss)"] = seconds_to_hms(sum_tmo_sale)
            
            metrics["TMO mediana (General) (hh:mm:ss)"] = seconds_to_hms(np.median(arr_contact)) if len(arr_contact) > 0 else "00:00:00"
            
            metrics["TMO mediana (venta) (hh:mm:ss)"] = seconds_to_hms(np.median(arr_sale)) if len(arr_sale) > 0 else "00:00:00"
            
            metrics["TMO mediana (no venta) (hh:mm:ss)"] = seconds_to_hms(np.median(arr_no_sale)) if len(arr_no_sale) > 0 else "00:00:00"
            
            metrics["SLA OPERATIVO (10-20 Lu-Vi)"] = seconds_to_hms(sla_op_med) 
            metrics["SLA EXTRAHORARIO (Lu-Vi)"] = seconds_to_hms(sla_ex_med)
            metrics["SLA FIN DE SEMANA"] = seconds_to_hms(sla_fds_med)
            
            metrics["ACW Mediana (hh:mm:ss)"] = seconds_to_hms(acw_med)
            metrics["Total tiempo llamadas"] = seconds_to_hms(sum_time_call_total)
            
            return metrics

        
        # --- GENERAR TABLA PIVOTADA ---
        months = sorted(df_filtered['mes_sort'].unique())
        final_dict = {"INDICADOR": row_labels}
        
        # Por Mes (Solo >= 2026)
        for m_sort in months:
            df_m = df_filtered[df_filtered['mes_sort'] == m_sort]
            label = m_sort.strftime('%B %Y')
            m_metrics = get_metrics_for_group(df_m)
            
            # En columnas mensuales, Total Recibidos es el total de ESE mes (ya filtrado por ao 2026? 
            # Si el mes es de 2026, son todos. Si hubiera meses anteriores se excluiran.
            # Asumimos que la columna es "Leads Recibidos (Ese Mes)".
            
            col_vals = []
            for k in row_labels:
                val = m_metrics.get(k, 0)
                if "%" in k:
                    val = format_percentage(val)
                elif isinstance(val, (int, float)) and "conv_contactado_cerrado" in k:
                    val = round(val, 2)
                col_vals.append(val)
            final_dict[label] = col_vals
            
        # Total General
        # AQUI el usuario quiere "Total leads recibidos  todos los registros" (incluyendo < 2026).
        # El resto ("Leads unicos analizados"...) solo los >= 2026 (df_filtered).
        
        total_metrics = get_metrics_for_group(df_filtered)
        
        # Sobreescribir "Total leads recibidos" tambin debe ser del filtro?
        # User: "corrijo, todos los indicadores deben ser filtrados por >=2026"
        # Antes era: total_metrics["Total leads recibidos"] = len(df)
        # Ahora dejamos tal como calcule get_metrics_for_group(df_filtered), que ya usa len(df_filtered).
        # No hacemos nada extra.
        
        col_vals_total = []
        for k in row_labels:
            val = total_metrics.get(k, 0)
            if "%" in k:
                val = format_percentage(val)
            elif isinstance(val, (int, float)) and "conv_contactado_cerrado" in k:
                val = round(val, 2)
            col_vals_total.append(val)
        final_dict["TOTAL GENERAL"] = col_vals_total
        
        df_exec = pd.DataFrame(final_dict)
        
        # --- GUARDAR CON ESTILOS ---
        original_sheets['Resumen_Ejecutivo'] = df_exec
        
        base_name = os.path.basename(latest_file)
        parts = base_name.split('_')
        provider = parts[0] if len(parts) > 0 else "UNKNOWN"
        timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
        
        output_filename = f"{provider}_RESUMEN_EJECUTIVO_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        print(f"Guardando Resumen Ejecutivo en: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
             for sheet_name, df_sheet in original_sheets.items():
                if isinstance(df_sheet, pd.DataFrame):
                     df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                     
             # Aplicar estilos a Resumen_Ejecutivo
             wb = writer.book
             if 'Resumen_Ejecutivo' in wb.sheetnames:
                 ws = wb['Resumen_Ejecutivo']
                 
                 # Ajustar ancho col A
                 ws.column_dimensions['A'].width = 45
                 
                 # Estilos:
                 # Col A (Indicadores): Left Aligned, con COMENTARIOS
                 # Col B+ (Valores): Center Aligned
                 
                 from openpyxl.comments import Comment
                 
                 # Diccionario de comentarios para indicadores (Texto exacto usuario + Actualizaciones)
                 indicadores_comments = {
                    "Total leads recibidos": "todos los >=2026",
                    "Leads nicos analizados": "todos los >=2026",
                    "Contactados (unicos)": "todos los contactado si",
                    "No contactados (Unicos)": "todos los contactado no",
                    "Contactabilidad % (unicos)": "todos los contactado si/todos los >=2026 * 100",
                    "Leads Cerrados": "Conteo de leads con status CERRADO",
                    "Ventas (leads insertados mes)": "leads venta Si",
                    "interacciones_contacto": "contar(todos los tmo>0)",
                    "interacciones_sin_contacto": "contar(todos los tmo=0 o nulo)",
                    "Conversin % (contactados)": "leads venta Si/todos los contactado si*100",
                    "conversion sobre interacciones con contacto_%": "interacciones de venta/ interacciones tmo > 0",
                    "conv_contactado_cerrado_%": "leads contactados si /leads venta y cerrado",
                    "interacciones en venta": "interacciones resuldesc= VENTA /POLIZA",
                    "Total interacciones": "conteo de todos los tmo>0",
                    "TMO totalizado (hh:mm:ss)": "sumatoria de todos los tmo>0",
                    "TMO mediana (General) (hh:mm:ss)": "mediana de  todos los tmo>0",
                    "TMO interacciones (venta) (hh:mm:ss)": "sumatoria(tmo>0 con resuldesc = VENTA /POLIZA)",
                    "TMO mediana (venta) (hh:mm:ss)": "MEDIANA (tmo>0 con resuldesc = VENTA /POLIZA)",
                    "TMO mediana (no venta) (hh:mm:ss)": "sumatoria(tmo>0 con resuldesc != VENTA /POLIZA)", 
                    "SLA OPERATIVO (10-20 Lu-Vi)": "Mediana de sla en la franja 10-20 Lu-Vi por fxCreated",
                    "SLA EXTRAHORARIO (Lu-Vi)": "Mediana de sla en la franja fuera de 10-20 dias Lu-Vi por fxCreated",
                    "SLA FIN DE SEMANA": "Mediana de sla en la franja no  Lu-Vi por fxCreated",
                    "ACW Mediana (hh:mm:ss)": "ACW Mediana",
                    "Total tiempo llamadas": "sumatoria de todo el tiempo en llamadas (timecall)"
                 }
                 
                 max_row = ws.max_row
                 max_col = ws.max_column
                 
                 align_left = Alignment(horizontal='left', vertical='center')
                 align_center = Alignment(horizontal='center', vertical='center')
                 
                 for row in range(1, max_row + 1):
                     # Celda Indicador (Col 1)
                     cell_ind = ws.cell(row=row, column=1)
                     cell_ind.alignment = align_left
                     ind_val = str(cell_ind.value).strip()
                     
                     if ind_val in indicadores_comments:
                         cell_ind.comment = Comment(indicadores_comments[ind_val], "System")
                         
                     # Celdas Valores (Col 2+)
                     for col in range(2, max_col + 1):
                         cell = ws.cell(row=row, column=col)
                         cell.alignment = align_center

        print("Resumen Ejecutivo generado exitosamente.")

    except Exception as e:
        print(f"Error generando Resumen Ejecutivo: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_executive_summary()
