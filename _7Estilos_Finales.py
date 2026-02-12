import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

# --- CONFIGURACION DE COLORES ---
# Azul Claro para Encabezados
COLOR_HEADER = "BDD7EE" 
# Coral para Totales (#FF8169)
COLOR_TOTAL = "FF8169"

# --- DICCIONARIOS DE COMENTARIOS ---
# Copiados de scripts anteriores (_5Resumen_Ejecutivo y _2Resumen_Diario)

COMMENTS_EJECUTIVO = {
    "Total leads recibidos": "todos los >=2026",
    "Leads Unicos analizados": "registros sin duplicado todos los >=2026",
    "Contactados (unicos)": "todos los contactado si",
    "No contactados (Unicos)": "todos los contactado no",
    "Contactabilidad % (unicos)": "todos los contactado si/todos los >=2026 * 100",
    "Leads Cerrados": "Conteo de leads con status CERRADO",
    "Ventas (leads insertados mes)": "leads venta Si",
    "interacciones_contacto": "contar(todos los tmo>0)",
    "interacciones_sin_contacto": "contar(todos los tmo=0 o nulo)",
    "Conversin % (contactados)": "leads venta Si/todos los contactado si*100",
    "conversion sobre interacciones con contacto_%": "interacciones de venta/ interacciones tmo > 0",
    "conv_contactado_cerrado_%": "leads venta /leads contactados si",
    "interacciones en venta": "interacciones resuldesc= VENTA /POLIZA",
    "Total interacciones": "conteo de todos los tmo>0",
    "TMO totalizado (hh:mm:ss)": "sumatoria de todos los tmo>0",
    "TMO mediana (General) (hh:mm:ss)": "mediana de  todos los tmo>0",
    "TMO interacciones (venta) (hh:mm:ss)": "sumatoria(tmo>0 con resuldesc = VENTA /POLIZA)",
    "TMO mediana (venta) (hh:mm:ss)": "MEDIANA (tmo>0 con resuldesc = VENTA /POLIZA)",
    "TMO mediana (no venta) (hh:mm:ss)": "sumatoria(tmo>0 con resuldesc != VENTA /POLIZA)", 
    "SLA OPERATIVO (10-20 Lu-Vi)": "Mediana de sla en la franja 10-20 Lu-Vi por fxCreated",
    "SLA EXTRAHORARIO (Lu-Vi)": "Mediana de sla en la franja fuera de 10-20 dias Lu-Vi por fxCreated",
    "SLA FIN DE SEMANA": "Mediana de sla en la franja no  Lu-Vi por fxCreated",
    "ACW Mediana (hh:mm:ss)": "ACW Mediana",
    "Total tiempo llamadas": "sumatoria de todo el tiempo en llamadas (timecall)"
}

COMMENTS_DIARIO_GENERIC = {
    # Columnas comunes y Agentes
    "fecha": "Fecha del registro",
    "agente": "Nombre del agente",
    "leads_insertados": "Total leads únicos",
    "contactados": "Leads con TMO>0",
    "ventas": "Leads con venta=SI y status=CERRADO",
    "leads_cerrados": "Leads únicos cerrados por el agente",
    "int_contacto": "Interacciones con TMO>0",
    "int_sin_contacto": "Interacciones con TMO=0",
    "interacciones_ventas": "Interacciones con Venta=SI",
    "interacciones_total": "Suma interacciones (contacto + sin contacto)",
    
    # Porcentajes Agentes y Diario
    "contactabilidad_%": "contactados / leads_insertados * 100",
    "conversion_%": "ventas / contactados * 100",
    "conversion_contactos_%": "interacciones_ventas / int_contacto * 100",
    "conv_contactado_cerrado_%": "ventas / leads_cerrados * 100",
    
    # Tiempos HMS
    "tiempo_total_llamadas_hms": "Suma total TimeCall",
    "tmo_total_hms": "Suma TMOs",
    "tmo_total_mediana_hms": "Mediana TMOs (General)",
    "tmo_ventas_hms": "Suma TMOs (Ventas)",
    "tmo_ventas_mediana_hms": "Mediana TMOs (Ventas)",
    "tmo_no_venta_hms": "Suma TMOs (No Ventas)",
    "tmo_no_venta_mediana_hms": "Mediana TMOs (No Ventas)",
    
    # SLA y ACW
    "sla_hms_medio": "Promedio SLA",
    "sla_mediana_hms": "Mediana SLA (Global)",
    "sla_operativo_mediana_hms": "Mediana SLA (10-18h L-V)",
    "sla_extra_mediana_hms": "Mediana SLA (Extra Horario)",
    "sla_fds_mediana_hms": "Mediana SLA (FDS)",
    "time_acw_hms": "Suma ACW",
    "acw_mediana_hms": "Mediana ACW",
    
    # Otros
    "mediana_tmo_x_periodo_hms": "Mediana TMO periodo",
    "hora_franja": "Franja horaria (fxCreated)"
}

# Mapeo Hoja -> Diccionario
SHEET_COMMENTS_MAP = {
    "Resumen_Ejecutivo": COMMENTS_EJECUTIVO,
    "Agentes": COMMENTS_DIARIO_GENERIC,
    "Resumen_Diario": COMMENTS_DIARIO_GENERIC,
    "Resumen_Semanal": COMMENTS_DIARIO_GENERIC,
    "Frecuencia": COMMENTS_DIARIO_GENERIC
}

def find_latest_file(directory, pattern='*.xlsx'):
    files = glob.glob(os.path.join(directory, pattern))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    if not files: return None
    return max(files, key=os.path.getmtime)

def apply_styles_and_order():
    # INPUT: Ahora toma de FRECUENCIA (salida del paso anterior)
    # Ruta base: ...\SMART CONECT\INFORMES\KPI_SMART\FRECUENCIA
    input_dir = r"C:\Users\dresdev\OneDrive\Desktop\SMART CONECT\INFORMES\KPI_SMART\FRECUENCIA"
    
    print(f"Buscando archivo para estilos en: {input_dir}")
    latest_file = find_latest_file(input_dir, pattern='*FRECUENCIA*.xlsx')
    
    if not latest_file:
        print("No se encontr archivo en FRECUENCIA.")
        return

    print(f"Procesando estilos en: {latest_file}")
    
    try:
        wb = load_workbook(latest_file)
        
        # --- 1. RENOMBRAR Y REORDENAR HOJAS ---
        # Renombrar "Resumen_Agentes" a "Agentes"
        sheet_names = wb.sheetnames
        agents_name = next((s for s in sheet_names if 'Resumen_Agentes' in s), None)
        if agents_name:
            wb[agents_name].title = "Agentes"
            print(f"Renombrado: {agents_name} -> Agentes")
        
        # Orden deseado: [Resumen_Ejecutivo, Agentes, ...resto en orden original]
        sheet_names = wb.sheetnames  # Actualizar lista tras renombrar
        
        exec_name = next((s for s in sheet_names if 'Resumen_Ejecutivo' in s), None)
        agents_final = next((s for s in sheet_names if s == 'Agentes'), None)
        
        # Logica mover
        # Mover Agentes primero al index 0
        if agents_final:
            idx = wb.sheetnames.index(agents_final)
            wb.move_sheet(wb[agents_final], offset=-idx)
            
        # Mover Ejecutivo despues al index 0 (empujando Agentes al 1)
        if exec_name:
            idx = wb.sheetnames.index(exec_name)
            wb.move_sheet(wb[exec_name], offset=-idx)
            
        print(f"Orden de hojas ajustado: {wb.sheetnames}")
        
        # --- 2. ESTILOS Y COMENTARIOS ---
        
        fill_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        fill_total = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Determinar diccionario de comentarios
            comments_dict = {}
            if sheet_name in SHEET_COMMENTS_MAP:
                comments_dict = SHEET_COMMENTS_MAP[sheet_name]
            else:
                s_lower = sheet_name.lower()
                if 'diario' in s_lower or 'frecuencia' in s_lower or 'agentes' in s_lower or 'semanal' in s_lower:
                    comments_dict = COMMENTS_DIARIO_GENERIC
            
            # Iterar Filas
            max_row = ws.max_row
            max_col = ws.max_column
            
            # CASO ESPECIAL: Resumen_Ejecutivo (comentarios en Col A, no en Fila 1)
            is_ejecutivo = sheet_name == "Resumen_Ejecutivo"
            
            for row_idx, row in enumerate(ws.iter_rows(), 1):
                # Check si es fila TOTAL (Columna 1 contiene 'TOTAL')
                is_total_row = False
                first_cell_val = str(row[0].value).upper() if row[0].value else ""
                
                if "TOTAL" in first_cell_val:
                    is_total_row = True
                
                for cell in row:
                    # Estilo Fila Total
                    if is_total_row:
                        cell.fill = fill_total
                        
                    # Encabezados (Fila 1)
                    if row_idx == 1:
                        cell.fill = fill_header
                        # Agregar Comentario (solo para hojas NO ejecutivo)
                        if not is_ejecutivo:
                            val = str(cell.value).strip() if cell.value else ""
                            if val in comments_dict:
                                if not cell.comment:
                                    cell.comment = Comment(comments_dict[val], "System")
                    
                    # COMENTARIOS ESPECIALES: Resumen_Ejecutivo (Columna A = Indicadores)
                    if is_ejecutivo and cell.column == 1 and row_idx > 1:
                        indicator_val = str(cell.value).strip() if cell.value else ""
                        if indicator_val in comments_dict:
                            if not cell.comment:
                                cell.comment = Comment(comments_dict[indicator_val], "System")
                    
                    # ALINEACION
                    # Columna A (idx 0 en iter_rows pero cell.column es 1-based)
                    # Si cell.column == 1 -> Left, sino Center
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                
        # --- 3. GUARDAR EN ENTREGABLES ---
        # Ruta Salida: ...\KPI_SMART\ENTREGABLES\DDMMYYYY_RESUMENES\R_{PROVIDER}_{DDMMYYYY}_{HHMMSS}.xlsx
        
        from datetime import datetime
        now = datetime.now()
        ddmmyyyy = now.strftime("%d%m%Y")
        hhmmss = now.strftime("%H%M%S")
        
        # Obtener Provider del nombre archivo input
        # Asumimos nombre tipo "PLAYFILM_FRECUENCIA_..."
        base_name = os.path.basename(latest_file)
        parts = base_name.split('_')
        provider = parts[0] if len(parts) > 0 else "UNKNOWN"
        
        # Directorio Base ENTREGABLES
        freq_dir = os.path.dirname(latest_file) # KPI_SMART/FRECUENCIA
        kpi_smart_dir = os.path.dirname(freq_dir) # KPI_SMART
        entregables_dir = os.path.join(kpi_smart_dir, "ENTREGABLES")
        
        final_dir = os.path.join(entregables_dir, f"{ddmmyyyy}_RESUMENES")
        
        if not os.path.exists(final_dir):
            os.makedirs(final_dir)
            
        final_filename = f"R_{provider}_{ddmmyyyy}_{hhmmss}.xlsx"
        final_path = os.path.join(final_dir, final_filename)
        
        wb.save(final_path)
        print(f"Estilos aplicados. Reporte FINAL guardado en: {final_path}")

    except Exception as e:
        print(f"Error aplicando estilos: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    apply_styles_and_order()
